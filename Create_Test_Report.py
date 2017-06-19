# the script is to create report automatically
# used package: openpyxl
# info is saved in the document: configuration.ini


import os,sys
import configparser
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors, Color, NamedStyle, fills
from fogbugz import FogBugz

class Report_Style:
    def __init__(self):
        pass

    def create_report_style(self):
        # set excel cell style
        self.BasicReportStyle = NamedStyle(name="ReportStyle")
        self.BasicReportStyle.font = Font(
            name='Calibri',
            size=15,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color=colors.BLACK
        )
        self.BasicReportStyle.alignment = Alignment(
            horizontal='centerContinuous',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0
        )
        self.BasicReportStyle.fill = PatternFill(
            fill_type=fills.FILL_SOLID,
            fgColor = '00B0C4DE'
        )
        self.BasicReportStyle.border = Border(
            top=Side(
                border_style= 'thin',
                color='FF000000'),
            left=Side(
                border_style= 'thin',
                color='FF000000'),
            right=Side(
                border_style= 'thin',
                color='FF000000'),
            bottom=Side(
                border_style= 'thin',
                color='FF000000')
        )

        # Title
        self.titleStyle = NamedStyle(name="Title")
        self.titleStyle = self.BasicReportStyle

        # sub title - bold
        self.boldtitleStyle = NamedStyle(name="boldTitle")
        self.boldtitleStyle.font = Font(
            name='Calibri',
            size=11,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color=colors.BLACK
        )
        self.boldtitleStyle.alignment = Alignment(
            horizontal='left',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0
        )
        self.boldtitleStyle.border = self.BasicReportStyle.border

        # sub content 1
        self.subTitleStyle1 = NamedStyle(name="SubTitle1")
        self.subTitleStyle1.font = Font(
            name='Calibri',
            size=11,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color=colors.BLACK
        )
        self.subTitleStyle1.fill = PatternFill(
            fill_type=fills.FILL_SOLID,
            fgColor = '00C0C0C0'
        )
        self.subTitleStyle1.alignment = self.BasicReportStyle.alignment
        self.subTitleStyle1.border = self.BasicReportStyle.border

        # sub content 2
        self.subTitleStyle2 = NamedStyle(name="SubTitle2")
        self.subTitleStyle2.font = Font(
            name='Calibri',
            size=11,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color=colors.BLACK
        )
        self.subTitleStyle2.fill = PatternFill(
            fill_type=fills.FILL_SOLID,
            fgColor = '00ADD8E6'
        )
        self.subTitleStyle2.alignment = Alignment(
            horizontal='left',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0
        )
        self.subTitleStyle2.border = self.BasicReportStyle.border

        # sub Title2 - center
        self.subTitleStyle2_center = NamedStyle(name="SubTitle2_center")
        self.subTitleStyle2_center.alignment = Alignment(
            horizontal='center',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0
        )
        self.subTitleStyle2_center.font = self.subTitleStyle2.font
        self.subTitleStyle2_center.border = self.subTitleStyle2.border
        self.subTitleStyle2_center.fill = self.subTitleStyle2.fill

        # sub content 1 - non bold
        self.contentStyle1 = NamedStyle(name="content1")
        self.contentStyle1.font = Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color=colors.BLACK
        )
        self.contentStyle1.alignment = self.BasicReportStyle.alignment
        self.contentStyle1.border = self.BasicReportStyle.border
        self.contentStyle1.number_format = 'yyyy-mm-dd'

        # sub content 2 - non bold
        self.contentStyle2 = NamedStyle(name="content2")
        self.contentStyle2.font = Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color=colors.BLACK
        )
        self.contentStyle2.alignment = Alignment(
            horizontal='left',
            vertical='center',
            text_rotation=0,
            wrap_text=False,
            shrink_to_fit=False,
            indent=0
        )
        self.contentStyle2.border = self.BasicReportStyle.border


        # content link
        self.contentLinkStyle = NamedStyle(name="contentlink")
        self.contentLinkStyle.font = Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='single',
            strike=False,
            color=colors.BLUE
        )
        self.contentLinkStyle.alignment = self.BasicReportStyle.alignment
        self.contentLinkStyle.border = self.BasicReportStyle.border



class Report_Creator:
    def __init__(self):
        pass

    # import test report info
    def get_report_info(self, file):
        if not os.path.exists(file):
            print("no report info file, please check again!")
            return None
        else:
            cf = configparser.ConfigParser()
            cf.read(file)
            sections = cf.sections()
            # get report info
            self.ReportInfoDict = {
                "Product" : cf.get('ReportInfo','Product'),
                "Branch" : cf.get('ReportInfo','Branch'),
                "RunType" : cf.get('ReportInfo','TestRunType'),
                "ReportType" : cf.get('ReportInfo','ReportType')
            }
            # get TestInfo
            self.TestInfoTitleDict = {
                1 : 'Start Time',
                2 : 'End Time',
                3 : 'Build ID',
                4 : 'Version ID',
                5 : 'Build Type',
                6 : 'Export Level'
            }
            self.TestInfoContentDict = {
                'Start Time' : cf.get('TestInfo','StartTime'),
                'Duration' : cf.get('TestInfo','Duration'),
                'End Time' : cf.get('TestInfo','EndTime'),
                'Build ID' : cf.get('TestInfo','BuildID'),
                'Version ID' : cf.get('TestInfo','VersionIDPath'),
                'Build Type' : cf.get('TestInfo','BuildType'),
                'Export Level' : cf.get('TestInfo','ExportLevel'),
                'TestRun ID' : cf.get('TestInfo','TestRunID'),
            }
            # get Test Contents
            self.Category = (cf.get('TestContents','Category')).split(',')  # it's a list
            self.TestOS_Dict = {}
            for i in ['WIN10', 'WIN7']:
                self.TestOS_Dict[i] = cf.get('TestContents','TestOS_'+i)
            self.Category_Dict = {} # it's a dict
            for i in range(len(self.Category)):
                self.Category_Dict[i] = {
                    "TestSuites" : cf.get('Category'+str(i),'TestSuites'),
                    "Tester" : cf.get('Category'+str(i),'Tester'),
                    "RunLink" : cf.get('Category'+str(i),'RunLink'),
                    "ResultsLink" : cf.get('Category'+str(i),'Results_Link'),
                }

            # get Results Analysis contents
            self.ResultsInfo = {
                "TestFilter" : cf.get('Results Analysis','Test_Filter'),
            }

            # get Bug Contents
            self.BugCategory = (cf.get('BugContents','Category')).split(',')  # it's a list
            self.BugCategoryNameDict = {
                'Graphics': {},
                'CUDA': {},
            }
            self.BugCategoryStatusDict = {
                'Graphics': {},
                'CUDA': {},
            }
            self.FogbugzDict = {
                'Logname': cf.get('FOGBUGZ','logname'),
                'Password':cf.get('FOGBUGZ','password'),
                'Link': cf.get('FOGBUGZ','Link'),
            }

    def GetBugInfo(self, file):
        StatusDict = {
            '1' : 'active',
            '2' : 'resolved(fixed)',
            '3' : 'closed(not reproducible)',
            '4' : 'closed(duplicate)',
            '5' : 'closed(postpond)',
            '6' : 'closed(won\'t fix)',
            '7' : 'closed(by design)',
            '8' : 'closed(Implemented)',
            '35' : 'closed(not a bug)',
        }
        logname = self.FogbugzDict['Logname']
        password = self.FogbugzDict['Password']
        fp = FogBugz(self.FogbugzDict['Link'])
        fp.logon(logname, password)
        fb = open('BugInfo.txt')
        lineIndex = iter(fb.readlines())
        pattern1 = '^([\d]+)[\s]+(.+)$'
        pattern1_1 = '([A-Za-z]+)'
        pattern2 = '([\d]+)[\s,]+'
        try:
            print('Begin to read the bug info ...')
            for line in lineIndex:
                if line[1:-2] in self.BugCategoryNameDict.keys():
                    Bugtag = line[1:-2]
                    tmpNameDict = {}
                    tmpStatusDict = {}
                    next(lineIndex)
                    continue
                if line != '\n' and line not in self.BugCategoryNameDict.keys():
                    match1 = re.match(pattern1,line)
                    match1_1 = re.search(pattern1_1,line)
                    match2 = re.findall(pattern2,line)
                    if match1 is not None and match1_1 is not None:
                        number = (match1.groups())[0]
                        resp = fp.search(q=number,cols='ixBug,sTitle,ixStatus')
                        tmpNameDict[number] = resp.cases.case.stitle.text
                        tmpStatusDict[number] = StatusDict[str(resp.cases.case.ixstatus.text)]
                        self.BugCategoryNameDict[Bugtag] = tmpNameDict
                        self.BugCategoryStatusDict[Bugtag] = tmpStatusDict
                    if match2 is not None and match1_1 is None:
                        for i in match2:
                            number = i
                            resp = fp.search(q=number,cols='ixBug,sTitle,ixStatus')
                            tmpNameDict[number] = resp.cases.case.stitle.text
                            tmpStatusDict[number] = StatusDict[str(resp.cases.case.ixstatus.text)]
                        self.BugCategoryNameDict[Bugtag] = tmpNameDict
                        self.BugCategoryStatusDict[Bugtag] = tmpStatusDict
        except StopIteration:
            print('Here is the end of the bug info file.')
        # print the bug list
        for cate_name in self.BugCategoryNameDict.keys():
            print(cate_name)
            for number in self.BugCategoryNameDict[cate_name]:
                print(number,self.BugCategoryNameDict[cate_name][number],self.BugCategoryStatusDict[cate_name][number])



    def Create_Excel(self, file, ReportStyle, ReportType):
        # enum
        GridDict = {
            1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G',
        }
        OSIndexDict = {
            1: 'WIN10', 2: 'WIN7',
        }
        Width = 4
        # call function get_report_info()
        self.get_report_info(file)

        # create excel
        wb = Workbook()
        ws = wb.active
        # load report styles
        wb.add_named_style(ReportStyle.titleStyle)

        # create an excel to record the result
        if ReportType == 'run' or (self.ReportInfoDict['ReportType']).lower() == 'run':
            ReportName = self.ReportInfoDict['Product']+' '+self.ReportInfoDict['Branch']+' '+self.ReportInfoDict['RunType']+' Run'
        elif ReportType == 'report' or (self.ReportInfoDict['ReportType']).lower() == 'report':
            ReportName = self.ReportInfoDict['Product']+' '+self.ReportInfoDict['Branch']+' '+self.ReportInfoDict['RunType']+' Test Report'
        else:
            ReportName = None
            print('Please indicate the report type!')
        ws.title = ReportName
        print('-------------------------------------------------------')
        print('The report name is '+ ReportName)
        print('-------------------------------------------------------')

        _rowIndex = 2
        _colIndex = 2
        # create title
        _row = _rowIndex
        _col =_colIndex
        title = ReportName
        ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width]+str(_row))
        _cell = ws[GridDict[_col]+str(_row)]
        _cell.value = title
        for j in range(_col,_col+Width+1):
            _cell = ws.cell(row=_row, column=j)
            _cell.style = ReportStyle.titleStyle

        _rowIndex = 3
        _colIndex = 2
        # create testinfo
        # Start Time, End Time, Build ID
        _row = _rowIndex
        _col = _colIndex
        for i in range(0, 4):
            _cell = ws.cell(row=_row+i, column=_col, value=self.TestInfoTitleDict[i+1])
            _cell.style = ReportStyle.boldtitleStyle
        _col = _colIndex + 1
        if self.TestInfoContentDict['Start Time'] == 'today':
            date_time = datetime.date.today()
            self.TestInfoContentDict['Start Time'] = date_time
        if self.TestInfoContentDict['End Time'] == 'N/A':
            date_time = self.TestInfoContentDict['Start Time'] + datetime.timedelta(days=int(self.TestInfoContentDict['Duration']))
            self.TestInfoContentDict['End Time'] = date_time
        for i in range(0,3):
            ws.merge_cells(GridDict[_col]+str(_row+i)+ ':' + GridDict[_col+Width-1]+str(_row+i))
            _cell = ws[GridDict[_col]+str(_row+i)]
            _cell.value = self.TestInfoContentDict[self.TestInfoTitleDict[i+1]]
            for j in range(_col,_col+Width):
                _cell = ws.cell(row=_row+i, column=j)
                _cell.style = ReportStyle.contentStyle1

        _rowIndex = 6
        _colIndex = 2
        # version ID
        _row = _rowIndex
        _col = _colIndex + 1
        ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width-1]+str(_row))
        _cell = ws[GridDict[_col]+str(_row)]
        _cell.value = ((self.TestInfoContentDict[self.TestInfoTitleDict[4]]).split('\\'))[-1]
        _cell.hyperlink = self.TestInfoContentDict[self.TestInfoTitleDict[4]]
        for j in range(_col,_col+Width):
            _cell = ws.cell(row=_row, column=j)
            _cell.style = ReportStyle.contentLinkStyle

        _rowIndex = 7
        _colIndex = 2
        # Build Type & Export Level
        _row = _rowIndex
        _col = _colIndex
        _cell = ws.cell(row=_row, column=_col, value=self.TestInfoTitleDict[5])
        _cell.style = ReportStyle.boldtitleStyle
        _col = _col + 1
        _cell = ws.cell(row=_row, column=_col, value=self.TestInfoContentDict[self.TestInfoTitleDict[5]])
        _cell.style = ReportStyle.contentStyle1
        _col = _col + 1
        _cell = ws.cell(row=_row, column=_col, value=self.TestInfoTitleDict[6])
        _cell.style = ReportStyle.boldtitleStyle
        _col = _col + 1
        ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width-3]+str(_row))
        _cell = ws[GridDict[_col]+str(_row)]
        _cell.value = self.TestInfoContentDict[self.TestInfoTitleDict[6]]
        for j in range(_col,_col+Width-2):
            _cell = ws.cell(row=_row, column=j)
            _cell.style = ReportStyle.contentStyle1

        _rowIndex = 8
        _colIndex = 2
        # test contents
        _row = _rowIndex
        _col = _colIndex
        ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width]+str(_row))
        _cell = ws[GridDict[_col]+str(_row)]
        _cell.value = 'Test Contents'
        for i in range(_col,_col+Width+1):
            _cell = ws.cell(row=_row, column=i)
            _cell.style = ReportStyle.subTitleStyle1
        # Category
        _row = _row + 1
        _col = 2
        _cell = ws.cell(row=_row, column=_col, value='Category')
        _cell.style = ReportStyle.subTitleStyle2
        for i in range(0,len(self.Category)):
            _cell = ws.cell(row=_row+1+i, column=_col, value=self.Category[i])
            _cell.style = ReportStyle.contentStyle2
        # Test Suites
        _row = _row
        _col = _col + 1
        _cell = ws.cell(row=_row, column=_col, value='Test Suites')
        _cell.style = ReportStyle.subTitleStyle2_center
        for i in range(0,len(self.Category)):
            _cell = ws.cell(row=_row+1+i, column=_col, value=(self.Category_Dict[i])['TestSuites'])
            _cell.style = ReportStyle.contentStyle2


        _rowIndex = 9
        _colIndex = 4
        # Run Link (the part is different in Run and Test Report)
        if ReportType == 'run' or (self.ReportInfoDict['ReportType']).lower() == 'run':
            # Tester
            _row = _rowIndex
            _col = _colIndex
            _cell = ws.cell(row=_row, column=_col, value='Tester')
            _cell.style = ReportStyle.subTitleStyle2
            for i in range(0,len(self.Category)):
                _cell = ws.cell(row=_row+1+i, column=_col, value=(self.Category_Dict[i])['Tester'])
                _cell.style = ReportStyle.contentStyle2
            # Test Link
            _row = _row
            _col = _col + 1
            ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width-3]+str(_row))
            _cell = ws[GridDict[_col]+str(_row)]
            _cell.value = 'Run Link'
            for i in range(_col,_col+Width-2):
                _cell = ws.cell(row=_row, column=i)
                _cell.style = ReportStyle.subTitleStyle2
            _row = _row + 1
            _col = _col
            for i in range(0,len(self.Category)):
                for j in range(0,2):
                    tmp = (self.Category_Dict[i])['RunLink'].replace('*TestRunID*', self.TestInfoContentDict['TestRun ID'])
                    Link = tmp.replace('*TestOSID*', self.TestOS_Dict[OSIndexDict[j+1]])
                    _cell = ws.cell(row=_row+i, column=_col+j, value='Link('+OSIndexDict[j+1]+')')
                    _cell.hyperlink = Link
                    _cell.style = ReportStyle.contentLinkStyle
        elif ReportType == 'report' or (self.ReportInfoDict['ReportType']).lower() == 'report':
            _row = _rowIndex
            _col = _colIndex
            ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width-2]+str(_row))
            _cell = ws[GridDict[_col]+str(_row)]
            _cell.value = 'Tester'
            for i in range(_col,_col+Width-1):
                _cell = ws.cell(row=_row, column=i)
                _cell.style = ReportStyle.subTitleStyle2_center
            for i in range(0,len(self.Category)):
                ws.merge_cells(GridDict[_col]+str(_row+1+i)+ ':' + GridDict[_col+Width-2]+str(_row+1+i))
                _cell = ws[GridDict[_col]+str(_row+1+i)]
                _cell.value = (self.Category_Dict[i])['Tester']
                for j in range(_col,_col+Width-1):
                    _cell = ws.cell(row=_row+1+i, column=j)
                    _cell.style = ReportStyle.contentStyle2
        else:
            print('Please indicate the report type!')

        _rowIndex = _rowIndex + 1 + len(self.Category)
        _colIndex = 2
        # Results Analysis
        _row = _rowIndex
        _col = _colIndex
        ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width]+str(_row))
        _cell = ws[GridDict[_col]+str(_row)]
        _cell.value = 'Results Analysis'
        for i in range(_col,_col+Width+1):
            _cell = ws.cell(row=_row, column=i)
            _cell.style = ReportStyle.subTitleStyle1
        # Category
        _row = _row + 1
        _col = 2
        _cell = ws.cell(row=_row, column=_col, value='Category')
        _cell.style = ReportStyle.subTitleStyle2
        for i in range(0,len(self.Category)):
            _cell = ws.cell(row=_row+1+i, column=_col, value=self.Category[i])
            _cell.style = ReportStyle.contentStyle2
        # Test Items
        _row = _row
        _col = _col + 1
        _cell = ws.cell(row=_row, column=_col, value='Test Items')
        _cell.style = ReportStyle.subTitleStyle2_center
        for i in range(0,len(self.Category)):
            _cell = ws.cell(row=_row+1+i, column=_col, value=self.ResultsInfo['TestFilter'])
            _cell.style = ReportStyle.contentStyle2
        # Percentage
        _row = _row
        _col = _col + 1
        _cell = ws.cell(row=_row, column=_col, value='Percentage')
        _cell.style = ReportStyle.subTitleStyle2
        for i in range(0,len(self.Category)):
            _cell = ws.cell(row=_row+1+i, column=_col, value='0%')
            _cell.style = ReportStyle.contentStyle2
        # Results Link
        _row = _row
        _col = _col + 1
        ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+Width-3]+str(_row))
        _cell = ws[GridDict[_col]+str(_row)]
        _cell.value = 'Results Link'
        for i in range(_col,_col+Width-2):
            _cell = ws.cell(row=_row, column=i)
            _cell.style = ReportStyle.subTitleStyle2_center
        _row = _row + 1
        _col = _col
        for i in range(0,len(self.Category)):
            ws.merge_cells(GridDict[_col]+str(_row+i)+ ':' + GridDict[_col+Width-3]+str(_row+i))
            _cell = ws[GridDict[_col]+str(_row+i)]
            _cell.value = 'Link'
            tmp = (self.Category_Dict[i])['ResultsLink']
            _cell.hyperlink = tmp.replace('*TestRunID*', self.TestInfoContentDict['TestRun ID'])
            for j in range(_col,_col+Width-2):
                _cell = ws.cell(row=_row+i, column=j)
                _cell.style = ReportStyle.contentLinkStyle

        _rowIndex = _rowIndex + 2 + len(self.Category)
        _colIndex = 2
        # create bug list
        if ReportType == 'report' or (self.ReportInfoDict['ReportType']).lower() == 'report':
            self.GetBugInfo('BugInfo.txt')
            for TestCate in self.Category:
                if bool(self.BugCategoryNameDict[TestCate].values()) is True:
                    # Bug list
                    _row = _rowIndex
                    _col = _colIndex
                    ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col+3]+str(_row))
                    _cell = ws[GridDict[_col]+str(_row)]
                    _cell.value = 'Bug List | ' + TestCate
                    for i in range(_col,_col+4):
                        _cell = ws.cell(row=_row, column=i)
                        _cell.style = ReportStyle.subTitleStyle2

                    # Status
                    _row = _rowIndex
                    _col = _colIndex + 4
                    ws.merge_cells(GridDict[_col]+str(_row)+ ':' + GridDict[_col]+str(_row))
                    _cell = ws[GridDict[_col]+str(_row)]
                    _cell.value = 'Status'
                    for i in range(_col,_col+1):
                        _cell = ws.cell(row=_row, column=i)
                        _cell.style = ReportStyle.subTitleStyle2_center

                    # write bugs
                    _row = _rowIndex + 1
                    _col = _colIndex
                    for t in sorted(self.BugCategoryNameDict[TestCate].keys()):
                        _cell = ws.cell(row=_row, column=_col, value=t)
                        _cell.hyperlink = self.FogbugzDict['Link']+ 'default.asp?'+t
                        _cell.style = ReportStyle.contentLinkStyle
                        ws.merge_cells(GridDict[_col+1]+str(_row)+ ':' + GridDict[_col+3]+str(_row))
                        _cell = ws[GridDict[_col+1]+str(_row)]
                        _cell.value = self.BugCategoryNameDict[TestCate][t]
                        for i in range(_col+1,_col+4):
                            _cell = ws.cell(row=_row, column=i)
                            _cell.style = ReportStyle.contentStyle2
                        ws.merge_cells(GridDict[_col+4]+str(_row)+ ':' + GridDict[_col+4]+str(_row))
                        _cell = ws[GridDict[_col+4]+str(_row)]
                        _cell.value = self.BugCategoryStatusDict[TestCate][t]
                        for i in range(_col+4,_col+5):
                            _cell = ws.cell(row=_row, column=i)
                            _cell.style = ReportStyle.contentStyle2
                        _row = _row + 1
                    _rowIndex = _rowIndex + len(self.BugCategoryNameDict[TestCate]) + 1

        # adjust the width of the table
        for column_cells in ws.columns:
            lengthList = []
            for cell in column_cells:
                if cell != ws['B2']:
                    lengthList.append(len(str(cell.value)))
                length = max(lengthList)
            ws.column_dimensions[column_cells[0].column].width = length

        # save the .xlsx file
        wb.save(ReportName+'__{'+self.TestInfoContentDict['Start Time']+'---'+self.TestInfoContentDict['End Time']+'}.xlsx')

    def Create_Email(self, file):
        pass

if __name__ == '__main__':
    if len(sys.argv) > 2:
        raise RuntimeError('You must enter only 2 parameters!')
    if len(sys.argv) == 2 and (sys.argv[1]).lower() != 'run' and (sys.argv[1]).lower() != 'report':
        raise RuntimeError('If there are 2 parameters, the second one must be run or report!')
    if len(sys.argv) == 2:
        ReportType = sys.argv[1]
    else:
        ReportType = False
    file = 'D:\Scripts_Work\Create_Test_Report\configuration.ini'
    ReportStyle = Report_Style()
    ReportStyle.create_report_style()
    report = Report_Creator()
    report.Create_Excel(file, ReportStyle, ReportType)
    print('done')